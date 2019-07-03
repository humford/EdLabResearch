import sys
import pprint
import openpyxl
import json

from googleapiclient.discovery import build
from fuzzywuzzy import fuzz

MAX_INDEX = 489596
STEP = 10000
OFFSET = 1

with open('cache.json', 'r') as f:
    cache = json.load(f)

# pprint.pprint(cache)

wb = openpyxl.load_workbook("research-sierra-test.xlsx")

read_sheet = wb["research-sierra"]
write_sheet = wb["test"]

api_key = "AIzaSyCAid0kMhZjasHpwIUjMWpM4h0uz4UP6tg"
service = build('books', 'v1', developerKey=api_key)

for i in range(2, MAX_INDEX, STEP):
	write_index = int(i / STEP) + OFFSET
	sierra_title = read_sheet.cell(row=i, column=6).value
	if (read_sheet.cell(row=i, column=2).value is not None):
		sierra_author = read_sheet.cell(row=i, column=2).value
	else:
		OFFSET = OFFSET - 1
		continue
		# sierra_title = ""
	sierra_id = read_sheet.cell(row=i, column=1).value
	# write_sheet.cell(row=i, column=1).value = sierra_title
	write_sheet["A" + str(write_index)] = sierra_title
	write_sheet["B" + str(write_index)] = sierra_author
	
	found = False
	cached = False

	try:
		response = cache[str(sierra_id)]
		if not response == None:
			found = True
		print(f"Found {sierra_id} in cache.")
		cached = True
	except KeyError:
		request = service.volumes().list(source='public', q=str(sierra_title + " " + sierra_author), maxResults=1)
		try:
			search = request.execute()
			if not search["totalItems"] == 0:
				response = search["items"][0]
				found = True
		except:
			print("API Problem found, skipping progress...")
			continue
	if found:
		if not cached:
			cache[sierra_id] = response
		write_sheet["C" + str(write_index)] = sierra_id
		try:
			google_title = response['volumeInfo']['title']
			write_sheet["D" + str(write_index)] = google_title
			write_sheet["E" + str(write_index)] = fuzz.ratio(sierra_title, google_title)
			write_sheet["F" + str(write_index)] = fuzz.partial_ratio(sierra_title, google_title)
		except KeyError:
			write_sheet["E" + str(write_index)] = 0
			write_sheet["F" + str(write_index)] = 0
		try:
			google_author = str(response['volumeInfo']['authors'][0])
			write_sheet["G" + str(write_index)] = google_author
			write_sheet["H" + str(write_index)] = fuzz.ratio(sierra_title, google_author)
			write_sheet["I" + str(write_index)] = fuzz.partial_ratio(sierra_title, google_author)
		except KeyError:
			write_sheet["H" + str(write_index)] = 0
			write_sheet["I" + str(write_index)] = 0
		try:
			write_sheet["L" + str(write_index)] = str(response['volumeInfo']['categories'][0])
		except KeyError:
			pass
		try:
			write_sheet["N" + str(write_index)] = response['searchInfo']['textSnippet']
		except KeyError:
			pass
	else:
		if not cached:
			cache[sierra_id] = None
	print(f"Wrote item {sierra_id} to spreadsheet")
	print(f"Finished {write_index} of {int(MAX_INDEX / STEP) + OFFSET}")

print("Saving results...")
wb.save("research-sierra-test.xlsx")
print("Results saved.")

print("Saving cache...")
with open('cache.json', 'w') as f:

    json.dump(cache, f)
print("Cache saved.")

#response["items"][0]["id"]
# pprint.pprint(response)

# print('Found %d books:' % len(response['items']))
# for book in response.get('items', []):
#   print('Title: %s' % (
#     book['volumeInfo']['title']))